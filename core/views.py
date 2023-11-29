import datetime

import pandas
import openpyxl
from django.contrib.auth.models import Group, Permission
from rest_framework import viewsets, views, generics, filters, permissions, response, mixins, decorators
from rest_framework.authentication import TokenAuthentication
from pharmacy.models import Produit
from utils import get_mongodb_client
from rest_framework.pagination import LimitOffsetPagination
from core import serializers

from core.models import Patient, Consultation, Constante, Service, RendezVous, Hospitalisation, \
    UniteHospitalisation, User, BoxHospitalisation



class DefaultPaginatorClass(LimitOffsetPagination):
    page_size = 20

class AuthUserMeView(views.APIView):
    permission_classes = (permissions.IsAuthenticated,)
    authentication_classes = [TokenAuthentication]

    def get(self, request, format=None):
        serializer = serializers.LoginResponseSerializer(request.user)
        return response.Response(serializer.data)




class PatientApiListView(generics.ListCreateAPIView):
    serializer_class = serializers.PatientSerializers
    queryset = Patient.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    pagination_class = DefaultPaginatorClass

    def get_queryset(self):
        code_patient = self.request.GET.get('code_patient', '').upper()
        nom = self.request.GET.get('nom', '').upper()
        prenoms = self.request.GET.get('prenoms', '').upper()
        contact = self.request.GET.get('contact', '').upper()
        return self.queryset.filter(
            code_patient__contains=code_patient,
            nom__contains=nom,
            prenoms__contains=prenoms,
            contact__contains=contact,
        )[:100]



class PatientApiDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = serializers.PatientDetailSerializers
    queryset = Patient.objects.all()


class ConsultationApiListView(generics.ListCreateAPIView):
    serializer_class = serializers.ConsultationSerializers
    queryset = Consultation.objects.all()


class UrgenceApiListView(generics.ListCreateAPIView):
    serializer_class = serializers.ConsultationSerializers
    queryset = Consultation.objects.filter(mode_entree='Urgence')


class ConsultationApiCreateView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = serializers.ConsultationSerializers
    queryset = Consultation.objects.all()


class ConstanteApiListView(generics.ListCreateAPIView):
    serializer_class = serializers.ConstanteSerializers
    queryset = Constante.objects.all()


class ServiceListView(generics.ListAPIView):
    serializer_class = serializers.ServiceSerializers
    queryset = Service.objects.all()


class ServiceConsultationListView(generics.ListAPIView):
    serializer_class = serializers.ConsultationSerializers

    def get_queryset(self):
        return Consultation.objects.filter(service_id=self.kwargs['service_id'])


class RendezVousListView(generics.ListCreateAPIView):
    serializer_class = serializers.RendezVousSerializer
    queryset = RendezVous.objects.filter(state=0)


class UniteHospitalisationListView(generics.ListCreateAPIView):
    serializer_class = serializers.UniteHospitalisationSerializer
    queryset = UniteHospitalisation.objects.all()


class HospitalisationListView(generics.ListCreateAPIView):
    serializer_class = serializers.HospitalisationSerializer
    queryset = Hospitalisation.objects.all()


class HospitalisationUpdateView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = serializers.HospitalisationSerializer
    queryset = Hospitalisation.objects.all()


class StatistiqueView(views.APIView):
    def get(self, request, format=None):
        return response.Response({
            'patients': Patient.objects.count(),
            'prestations': Consultation.objects.count(),
            'personels': User.objects.count(),
            'Produits': Produit.objects.count(),
        })


class BilanInitialListView(views.APIView):
    def post(self, request, format=None):
        db = get_mongodb_client()
        data = request.data
        data['created_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        db['bilan_initial'].insert_one(data)
        return response.Response({'message': 'success'})

    def get(self, request, pk, format=None):
        db = get_mongodb_client()
        # get patient id in path param

        return response.Response(db['bilan_initial'].find({'patient_id': pk}))


class DossierDataAPIView(views.APIView):

    def get(self, request, pk, collection):
        db = get_mongodb_client()
        data = db[collection].find_one({'patient_id': pk})

        # todo: make object_id json serializable
        return response.Response(data)

    def post(self, request, collection):
        db = get_mongodb_client()
        data = request.data
        data['user_id'] = self.request.user.pk
        db[collection].insert_one(data)

    def update(self, request, collection, pk):
        db = get_mongodb_client()
        data = request.data
        db[collection].update_one({'patient_id': pk}, data)


class BilanInitialView(views.APIView):
    pass


def get_rows_as_dict(worksheet):
    rows_as_dict = []

    # Assuming the first row contains the column headers.
    headers = [cell for cell in next(worksheet.iter_rows(values_only=True))]

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        rows_as_dict.append(row_dict)

    return rows_as_dict


def create_patient_fromexcel(file):
    rows = pandas.read_excel(file)
    print(rows.get('IDENT'))


def upload_excel_documents(file):
    print('start upload excel documents')
    sheet = openpyxl.load_workbook(file).active
    get_rows = get_rows_as_dict(sheet)
    print(get_rows[0])

    db = get_mongodb_client()
    for p in get_rows:
        try:
            patient = Patient.objects.create(
                code_patient=p['IDENT'],
                nom=p['NOM'] if p['NOM'] is not None else "",
                prenoms=p['PRENOM'] if p['PRENOM'] is not None else "",
                genre=p['SEXE'] if p['SEXE'] is not None else "",
                date_naissance=p['DATENAIS'] if p['DATENAIS'] is not None else "",
                lieu_naissance=p['VILLE'] if p['VILLE'] is not None else "",
                situation_matrimoniale=p['SITMAT'] if p['SITMAT'] is not None else "",
                niveau_etude=p['NIVETU'],
                nationalite=p['NATIONAL'] if p['NATIONAL'] is not None else "",
                contact=p['TELEPHONE'] if p['TELEPHONE'] is not None else ""
            )
            patient.domiciles.create(ville=p['VILLE'],
                                     commune=p['COMMUNE'])
        except:
            pass
        finally:
            db['patient_suivi'].insert_one(p)
    print('ending upload excel documents')





@decorators.api_view()
def get_boxes(request):
    boxes = BoxHospitalisation.objects.filter(occuper=False)

    return response.Response(
        serializers.BoxHospitalisationSerializer(boxes, many=True).data
    )


@decorators.api_view()
def suivre_patient(request, pk):
    patient = Patient.objects.get(pk=pk)
    patient.status = 1
    patient.save()
    return response.Response({'ok': True})


class UserListView(generics.ListCreateAPIView):
    serializer_class = serializers.UserSerializer
    queryset = User.objects.exclude(is_superuser=True)




class UserRetrieveView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = serializers.UserSerializer
    queryset = User.objects.exclude(is_superuser=True)


class GroupListView(generics.ListCreateAPIView):
    serializer_class = serializers.GroupSerializer
    queryset = Group.objects.all()


class PermissionListView(generics.ListCreateAPIView):
    serializer_class = serializers.PermissionSerializer
    queryset = Permission.objects.all()


class GroupViewset(viewsets.ModelViewSet):
    serializer_class = serializers.GroupSerializer
    queryset = Group.objects.all()
